"""Menu import services for CSV/Excel files."""

import csv
import io
import logging
from abc import ABC, abstractmethod
from typing import BinaryIO, Optional

logger = logging.getLogger(__name__)


class MenuImporter(ABC):
    """Base class for menu importers."""

    REQUIRED_COLUMNS = ["name", "price"]
    OPTIONAL_COLUMNS = [
        "description",
        "category",
        "is_available",
        "allergens",
        "dietary_tags",
        "spice_level",
        "prep_time_minutes",
        "ingredients",
    ]

    @abstractmethod
    def parse(self, file: BinaryIO) -> list[dict]:
        """
        Parse a file and return list of menu items.

        Args:
            file: File-like object to parse

        Returns:
            List of dicts with menu item data
        """
        pass

    def validate_row(self, row: dict, row_num: int) -> tuple[dict, list[str]]:
        """
        Validate a parsed row and return cleaned data.

        Args:
            row: Dict with row data
            row_num: Row number for error messages

        Returns:
            Tuple of (cleaned_data, errors)
        """
        errors = []
        cleaned = {}

        # Required fields
        name = row.get("name", "").strip()
        if not name:
            errors.append(f"Row {row_num}: Missing required field 'name'")
        else:
            cleaned["name"] = name

        price_str = str(row.get("price", "")).strip()
        if not price_str:
            errors.append(f"Row {row_num}: Missing required field 'price'")
        else:
            try:
                # Handle comma as decimal separator
                price_str = price_str.replace(",", "").replace(" ", "")
                cleaned["price"] = float(price_str)
            except ValueError:
                errors.append(
                    f"Row {row_num}: Invalid price '{row.get('price')}'"
                )

        # Optional fields
        if "description" in row and row["description"]:
            cleaned["description"] = str(row["description"]).strip()

        if "category" in row and row["category"]:
            cleaned["category"] = str(row["category"]).strip()

        if "is_available" in row:
            val = str(row["is_available"]).lower().strip()
            cleaned["is_available"] = val in ("true", "yes", "1", "oui")

        # Array fields (comma-separated)
        if "allergens" in row and row["allergens"]:
            allergens = [
                a.strip().lower()
                for a in str(row["allergens"]).split(",")
                if a.strip()
            ]
            cleaned["allergens"] = allergens

        if "dietary_tags" in row and row["dietary_tags"]:
            tags = [
                t.strip().lower()
                for t in str(row["dietary_tags"]).split(",")
                if t.strip()
            ]
            cleaned["dietary_tags"] = tags

        # Numeric fields
        if "spice_level" in row and row["spice_level"]:
            try:
                level = int(row["spice_level"])
                if 0 <= level <= 4:
                    cleaned["spice_level"] = level
            except ValueError:
                pass

        if "prep_time_minutes" in row and row["prep_time_minutes"]:
            try:
                cleaned["prep_time_minutes"] = int(row["prep_time_minutes"])
            except ValueError:
                pass

        if "ingredients" in row and row["ingredients"]:
            cleaned["ingredients"] = str(row["ingredients"]).strip()

        return cleaned, errors


class CSVImporter(MenuImporter):
    """Import menu items from CSV files."""

    def parse(self, file: BinaryIO) -> list[dict]:
        """Parse CSV file and return menu items."""
        items = []
        errors = []

        try:
            # Try to detect encoding
            content = file.read()
            try:
                text = content.decode("utf-8")
            except UnicodeDecodeError:
                text = content.decode("latin-1")

            # Reset for potential re-read
            reader = csv.DictReader(io.StringIO(text))

            # Normalize column headers
            if reader.fieldnames:
                fieldnames = [f.lower().strip() for f in reader.fieldnames]
                reader.fieldnames = fieldnames

            for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is 1)
                # Normalize keys
                normalized_row = {k.lower().strip(): v for k, v in row.items()}
                cleaned, row_errors = self.validate_row(normalized_row, row_num)

                if row_errors:
                    errors.extend(row_errors)
                elif cleaned:
                    items.append(cleaned)

        except Exception as e:
            logger.error(f"CSV parsing error: {e}")
            errors.append(f"Failed to parse CSV: {str(e)}")

        return {"items": items, "errors": errors}


class ExcelImporter(MenuImporter):
    """Import menu items from Excel files."""

    def parse(self, file: BinaryIO) -> list[dict]:
        """Parse Excel file and return menu items."""
        items = []
        errors = []

        try:
            import openpyxl

            # Read workbook
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active

            if not sheet or sheet.max_row < 2:
                return {"items": [], "errors": ["Empty or invalid Excel file"]}

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower().strip())
                else:
                    headers.append("")

            # Process data rows
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = value

                cleaned, row_errors = self.validate_row(row_dict, row_num)

                if row_errors:
                    errors.extend(row_errors)
                elif cleaned:
                    items.append(cleaned)

        except ImportError:
            errors.append("openpyxl package not installed. Run: pip install openpyxl")
        except Exception as e:
            logger.error(f"Excel parsing error: {e}")
            errors.append(f"Failed to parse Excel: {str(e)}")

        return {"items": items, "errors": errors}


def get_importer_for_file(filename: str) -> Optional[MenuImporter]:
    """
    Get the appropriate importer based on file extension.

    Args:
        filename: Name of the file

    Returns:
        MenuImporter instance or None if unsupported
    """
    ext = filename.lower().split(".")[-1] if "." in filename else ""

    if ext == "csv":
        return CSVImporter()
    elif ext in ("xlsx", "xls"):
        return ExcelImporter()
    return None
